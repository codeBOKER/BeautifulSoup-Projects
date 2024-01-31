"""
This script extracts product information from The Cycle Division website and saves it to an excel file.

Usage:
1. Install the required libraries by running the following command in your terminal: pip install requests beautifulsoup4 openpyxl
2. Save the code to a python file and run it by typing python <filename>.py in your terminal

Note: This script assumes that you have a directory called "data" in the same directory as the script, and that the directory contains an excel file called "products.xlsx"
"""

import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup


WW = Workbook()
sheets = WW.active


for page in range(1, 12):
    url = f"https://www.thecycledivision.com/product-category/workshop/page/{page}/"
    reaslut = requests.get(url)
    soup = BeautifulSoup(reaslut.content, "html.parser")
    
    # for every page a new sheeet
    sheets = WW.create_sheet(f"sheet{page}")

    # main parts
    parts = soup.find_all(class_="container-inner")
    for part in parts:
        # _______product_name_______
        name = part.find(class_="product-name").text

        # _______product_sku_______
        sku = part.find(class_="sku").text
        
        # _______product_link______
        link = part.find(class_="woocommerce-LoopProduct-link woocommerce-loop-product__link").get("href")

        # put things in excel file
        sheets.cell(column= 1 , row= parts.index(part) +2 , value= name)
        sheets.cell(column= 2 , row= parts.index(part) +2 , value= sku)
        sheets.cell(column= 3 , row= parts.index(part) +2 , value= link)

        WW.save("C:\\Users\\ab3ad\\Desktop\\cycle.xlsx")

#codeBOKAR