"""
Import Food packaging from iqsdirectory.com site :)
"""

import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup


WW = Workbook()
sheets = WW.active

reaslut = requests.get("https://www.iqsdirectory.com/food-packaging/food-packaging-2/")
soup = BeautifulSoup(reaslut.content, "html.parser")

# len(soup.find_all(class_="addr"))
for i in range(7):
    # _____campany_name_____
    f_ca_name = soup.find_all(class_="addr")[i].find("span").text
    ca_name = f_ca_name[:f_ca_name.index(",")]

    # ____campany_website____
    ca_web = soup.find_all(class_="cname")[i].find("a").get("href")

    # ____description____
    description = soup.find_all(class_="cdesc")[i].text
    

    sheets.cell(column= 1 , row= i+2 , value= ca_name)
    sheets.cell(column= 2 , row= i+2 , value= ca_web)
    sheets.cell(column= 3 , row= i+2 , value= description)
    WW.save("C:\\Users\\ab3ad\\Desktop\\projects_upwork\\food packaing\\food_packaing.xlsx")

#codeBOKER
    
