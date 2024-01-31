"""
Our code will grap (descriptions - price - link)
of the WIFIs in the -newegg- website to excel sheets
"""

import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
from urllib.request import urlopen

WW = Workbook()
sheets = WW.active

for page in range(1, 19):
    url = f"https://www.newegg.com/p/pl?d=at+t+wifi+router&Order=3&page={page}"
    reaslut = requests.get(url)
    soup = BeautifulSoup(reaslut.content, "lxml")
    
    # creat sheets
    sheets = WW.create_sheet(f"sheet_{page}")

    # limit what we will search about
    parts = soup.find_all(class_="item-cell")
    
    for part in parts:
        # _______discrption_______
        discrption = part.find(class_="item-title").text
        
        # ________price________
        try:
            f_price = str(part.find(class_="price-current").find("strong").text)
            s_price = str(part.find(class_="price-current").find("sup").text)
            price = f_price + s_price
        except AttributeError:
            f_price = str(part.find(class_="price-was-data").text)
            price = f_price + " changed"

        # _________link_________
        f_link = part.find(class_="item-img").get("href")
        link = f_link[0:f_link.index(" ")]


        # put the data in excel file 
        
        sheets.cell(column= 1 , row= parts.index(part)+ 2, value= discrption)
        sheets.cell(column= 2 , row= parts.index(part)+ 2, value= price)
        sheets.cell(column= 3 , row= parts.index(part)+ 2, value= link)
        WW.save("C://Users//ab3ad//Desktop//newegg1//wifi.xlsx")
        
#codeBOKER